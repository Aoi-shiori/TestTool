# -*- coding: utf-8 -*-
"""
# @Creation time: 2025/1/15 23:34
# @Author       : 郭军
# @Email        : 391350540@qq.com
# @FileName     : AWS_S3_File_Download.py
# @Software     : PyCharm
# @PythonVersion: python 3.12
# @Version      : 
# @Description  : 
# @Update Time  : 
# @UpdateContent:  

"""

import boto3
import openpyxl


def get_fileList():
    all_down_List={}
    workbook = openpyxl.load_workbook(config["excel"])
    sheetname = workbook.sheetnames
    for i in sheetname:
        downList_rawDataUrl = []
        downList_annotationsUrl = []
        # print(111999,i)
        sheet = workbook[i]
        # print(sheet.max_row)

        # # 遍历从第二行开始的所有行（忽略首行）
        # for k, row in enumerate(sheet.iter_rows(min_row=2, max_row=100000)):
        #     # row 是一个包含该行所有单元格对象的元组
        #     # linejson在左，xml在右,读取第一列和第二列的值
        #     # print(111,row)
        #     # 提取第一列和第二列的值
        #     # rawDataUrl,annotationsUrl = row[:2]
        #     rawDataUrl,annotationsUrl = row[:2]
        #     rawDataUrl=rawDataUrl.value
        #     annotationsUrl=annotationsUrl.value
        #     # print(k)
        #     if rawDataUrl is None:
        #         pass
        #     else:
        #         # 截取link中的文件名(rawDataUrl,annotationsUrl)
        #         _rawDataUrl=rawDataUrl.split("//")[1].split("/",maxsplit=1)[1]
        #         # print("数据行号：",k,"文件名：",_rawDataUrl)
        #         downList_rawDataUrl.append(_rawDataUrl)
        #
        #     if annotationsUrl is None:
        #         pass
        #     else:
        #         _annotationsUrl=annotationsUrl.split("//")[1].split("/",maxsplit=1)[1]
        #         # print("数据行号：",k,"文件名：",_annotationsUrl)
        #         downList_annotationsUrl.append(_annotationsUrl)
        # # print(downList)

        for row_index, row_data in get_data_by_column_names(sheet, ['rawDataUrl', 'annotationsUrl']):
            rawDataUrl = row_data['rawDataUrl']
            annotationsUrl = row_data['annotationsUrl']

            if rawDataUrl is None:
                continue
            else:
                # 截取link中的文件名(rawDataUrl,annotationsUrl)
                _rawDataUrl=rawDataUrl.split("//")[1].split("/",maxsplit=1)[1]
                # print("数据行号：",k,"文件名：",_rawDataUrl)
                downList_rawDataUrl.append(_rawDataUrl)

            if annotationsUrl is None:
                pass
            else:
                _annotationsUrl=annotationsUrl.split("//")[1].split("/",maxsplit=1)[1]
                # print("数据行号：",k,"文件名：",_annotationsUrl)
                downList_annotationsUrl.append(_annotationsUrl)


        # for row in sheet.iter_rows(min_row=2, values_only=True,max_row=100000):
        #     # row 是一个包含该行所有单元格值的元组
        #     # linejson在左，xml在右,读取第一列和第二列的值
        #     rawDataUrl,annotationsUrl = row[:2]
        #     # print(rawDataUrl,annotationsUrl)
        #     if rawDataUrl is None or annotationsUrl is None:
        #         pass
        #     else:
        #         # 截取link中的文件名(rawDataUrl,annotationsUrl)
        #         _rawDataUrl=rawDataUrl.split("//")[1].split("/",maxsplit=1)[1]
        #         downList.append(_rawDataUrl)
        #
        #         _annotationsUrl=annotationsUrl.split("//")[1].split("/",maxsplit=1)[1]
        #         # downList.append(_annotationsUrl)
        print("sheetname:",i,"\nawDataUrl总量:",len(downList_rawDataUrl),"\nannotationsUrl总量:",len(downList_annotationsUrl),"\n合计数量:",len(downList_rawDataUrl)+len(downList_annotationsUrl))
        write_log(f"sheetname:{i} \nawDataUrl总量:{len(downList_rawDataUrl)} \nannotationsUrl总量:{len(downList_annotationsUrl)} \n合计数量:{len(downList_rawDataUrl)+len(downList_annotationsUrl)}")
        all_down_List[i]=downList_rawDataUrl+downList_annotationsUrl

    return all_down_List


def get_data_by_column_names(sheet, target_columns):
    """
    根据列名获取数据

    Args:
        sheet: openpyxl worksheet对象
        target_columns: 目标列名列表，如 ['rawDataUrl', 'annotationsUrl']

    Returns:
        generator: 生成每行数据的字典
    """
    # 获取表头
    headers = [cell.value for cell in sheet[1]]

    # 获取目标列的索引
    column_indices = {}
    for col_name in target_columns:
        if col_name in headers:
            column_indices[col_name] = headers.index(col_name)
        else:
            logger.info(f"警告: 未找到列 '{col_name}'")
            return

    # 遍历数据行
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row_data = {col_name: row[idx] for col_name, idx in column_indices.items()}
        yield row_index, row_data


def downloadfile():
    file_list = []
    # # 孟买 041972482616
    # session = boto3.Session(
    #     aws_access_key_id="",
    #     aws_secret_access_key="",
    # )

    # 巴黎 270011752874
    session = boto3.Session(
        aws_access_key_id="",
        aws_secret_access_key="",
    )
    s3 = session.resource('s3')
    bucket = s3.Bucket(config['basurl'])

    # for obj in bucket.objects.all():
    #     file = obj.key
    #     print(file)
    # file_path = f'./heartcor/{file.split("/")[-1]}'
        # if "ECGRec_202210_C870258" in obj.key and "denoised.ecg" in obj.key:
        #     print(obj.key)
    # bucket.download_file(
    #     "2024-08-14/ECGRec_202426/C984244/1723645625556-1723645786556_66bcbae9729aa3dc13d434f8_denoised.xml",
    #     "./1723645625556-1723645786556_66bcbae9729aa3dc13d434f8_denoised.xml")

    downList=get_fileList()

    for k,v in downList.items():
        # print(656565,k,v)
        # for i in range(0,len(v),2):
        print(f"sheetname:{k} 开始下载！")
        for i in v:

            filname = f"./download/{k}/"+i.split("/")[-1]
            print(f"正在下载文件... sheetname：{k} --> {v.index(i)}:",filname)
            write_log(f"正在下载文件... sheetname：{k} --> {v.index(i)}:{filname}")

            bucket.download_file(i, filname)
        print(f"sheetname:{k} 下载完成！")

    print("Done")
            # file_name = "_".join(obj.key.split("/")[-1].split("_")[0:-1])
    #         S3_URI = f"s3://genentech-vitals-archive/{obj.key}"
    #         print(S3_URI)
    #         file_list.append({"file_name": file_name, "s3_url": S3_URI})
    # return file_list

# 记录日志
def write_log(log):
    with open("log.txt", "a") as f:
        f.write(log + "\n")
        f.close()



if __name__ == '__main__':
    # config={
    #     "basurl" : "webportal-us-west.vitals",
    #     "excel" : "medepace_patient_data.xlsx"
    # }
    config={
        "basurl" : "webportal-eu-west-3.vitals",
        "excel" : "./数据文件夹/ecgDataChunks_20251104.xlsx"
    }
    downloadfile()
    # get_fileList()
    # list=get_fileList()

