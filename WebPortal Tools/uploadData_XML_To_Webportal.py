#!/usr/bin/env python
# -*- coding: UTF-8 -*-
'''
@Time    : 2024-12-29 15:28:16
@File    : uploadData_XML.py
@Author  : Jun
@Software: PyCharm
@Description: 上传数据到webportal
'''
import json

import requests
import openpyxl
import re
import time
from concurrent.futures import ThreadPoolExecutor
from logger import logger
import  random


def Sendto_webportal(rawDataUrl, annotationsUrl, start, finish):
    startTime = int(start)
    finishTime = int(finish)
    if annotationsUrl is None:
        annotationsUrl = ""
    # 推送数据
    url = f"https://{config["baseurl"]}/api/backend/ecgDownloadTasks"
    # 重推数据
    # url = f"https://{config["baseurl"]}/api/backend/redownloadTasks"

    headers = {
        'Content-Type': 'application/json',
    }
    jsons = {
    "type":"EcgRaw",
    "deviceId": config["deviceId"],
    "start": startTime,
    "finish": finishTime,
    "rawDataUrl": rawDataUrl,
    "annotationsUrl": annotationsUrl,
    "sessionId": config["sessionId"]
    }
    payload = json.dumps(jsons)
    # print("发送数据：",payload,end="\t")
    response = requests.post(url, data=payload, headers=headers)
    return response

# 从rawDataUrl中提取开始时间、结束时间和患者ID
def get_file_name_info(filename):
    rawDataUrl=filename

    # 分割字符串以获取文件名部分
    # 示例：s3://webportal-test.vitals/2024-12-25/ECGRec_202420/DEVTEST13/1735095142548-1735095341884_676a1a4724db001eaf09880b_denoised.linejson.gz
    filename_with_ext = rawDataUrl.split('/')[-1]  # 获取最后一个'/'后的部分, 即完整文件名：1734518036712-1734518391282_671dbfc601d12f522af2660f.linejson.gz

    filename, ext = filename_with_ext.rsplit('.', 1)  #从右边开始，以'.'分割字符串，最多分割一次, 获取文件名和扩展名,并且扩展名不包含'.'。如果文件名中包含'.'，请使用maxsplit参数。文件名：1734518036712-1734518391282_671dbfc601d12f522af2660f.linejson，扩展名：gz
    # print(filename, ext)

    # 根据文件名和扩展名之间的'_'分割字符串
    parts = filename.split('_')
    # print(9999,parts)
    patient_id = parts[1]# 获取最后一个'_'后的部分, 即患者ID：671dbfc601d12f522af2660f
    try:
        startTime,endTime= parts[0].split('-') # 获取第一个'-'前的部分, 即开始时间戳：1734518036712，获取第一个'-'后的部分, 即结束时间戳：1734518391282
    except:
        logger.info(f'文件名格式错误！')
    # print("提取的的文件名数据：",startTime,endTime,patient_id)
    return startTime,endTime,patient_id

# 读取excel文件,生成上传数据
def get_UP_list(_sheetname):
    # config = {
    #     "excle": "ecgDataChunks_66bda9b22acb763421454137_UAT.xlsx",
    # }

    workbook = openpyxl.load_workbook(config["excle"])
    sheetname = workbook.sheetnames
    # print(len(sheetname))
    if _sheetname not in sheetname:
        logger.info(f'表格中没有找到sheetname：{_sheetname}')
        return None
    else:
        logger.info(f'找到sheetname：{_sheetname}')
        sheet = workbook[_sheetname]
        UP_LIST = {}
        # # 遍历从第二行开始的所有行（忽略首行）
        # for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        #     # row 是一个包含该行所有单元格值的元组
        #     # linejson在左，xml在右
        #     # rawDataUrl, annotationsUrl = row[:2]
        #     # 获取指定列Q,R列，第二行开始的数据
        #     rawDataUrl, annotationsUrl = row[16], row[17]
        #     # rawDataUrl, annotationsUrl = row[8], row[9]
        #     if rawDataUrl is None:
        #         pass
        #     # print(rawDataUrl)
        #     # print(annotationsUrl)
        #
        #     startTime, endTime, patient_id = get_file_name_info(rawDataUrl)
        #
        #     # 保存数据
        #     UP_LIST[row_index] = (rawDataUrl, annotationsUrl, startTime, endTime)
        # logger.info(type(UP_LIST))

        for row_index, row_data in get_data_by_column_names(sheet, ['rawDataUrl', 'annotationsUrl']):
            rawDataUrl = row_data['rawDataUrl']
            annotationsUrl = row_data['annotationsUrl']

            if rawDataUrl is None:
                continue

            startTime, endTime, patient_id = get_file_name_info(rawDataUrl)
            UP_LIST[row_index] = (rawDataUrl, annotationsUrl, startTime, endTime)
        logger.info(type(UP_LIST))
        return UP_LIST


def get_data_by_column_names(sheet, target_columns):
    """
    根据列名获取数据

    Args:
        sheet: openpyxl worksheet对象
        target_columns: 目标列名列表，如 ['rawDataUrl', 'annotationsUrl']

    Returns:
        generator: 生成每行数据的字典
    """
    # 获取表头
    headers = [cell.value for cell in sheet[1]]

    # 获取目标列的索引
    column_indices = {}
    for col_name in target_columns:
        if col_name in headers:
            column_indices[col_name] = headers.index(col_name)
        else:
            logger.info(f"警告: 未找到列 '{col_name}'")
            return

    # 遍历数据行
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row_data = {col_name: row[idx] for col_name, idx in column_indices.items()}
        yield row_index, row_data

# 写入日志
def write_log(log):
    # 使用utf-8编码
    with open("uploadData_XML.log", "a",encoding="UTF-8") as f:
        f.write(log)

import time

# 主函数
def main(config):
    config = config
    UP_LIST = get_UP_list(config["sheetname"])
    if UP_LIST is None:
        logger.info(f'上传数据为空,请检查表格数据！')
        exit()
    else:
        logger.info(f'开始数据上传时间：{time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())}\n')
        # write_log(f"开始数据上传时间：{time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())}\n")
        # 遍历数据
        for k, v in UP_LIST.items():
            rawDataUrl, annotationsUrl, startTime, endTime = v
            # 打印开始时间
            logger.info(f'{time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())}：正在上传数据...{config["env"]},sheetname：{config["sheetname"]}  表格数据行号：{k} --> {v}\n')

            max_retries = 5
            for attempt in range(max_retries):
                try:

                    response = Sendto_webportal(rawDataUrl, annotationsUrl, startTime, endTime)
                    if response.status_code == 200:
                        logger.info(f'数据上传成功！{response.status_code} , {response.content}\n')
                        logger.info(f'{"*"*100}')
                        break
                    else:
                        raise Exception(f"数据上传出错！{response.status_code} , {response.content}\n")
                except Exception as e:
                    logger.info(f"数据上传出错！ 尝试 {attempt + 1}/{max_retries} {e}\n")
                    if attempt < max_retries - 1:
                        time.sleep(5)  # 等待5秒后重试
                    else:
                        logger.info(f'数据上传失败，已达到最大重试次数。\n')
                        logger.info(f'{"*"*100}')

        logger.info(f'完成数据上传时间：{time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())}\n')

# 主函数
def main(config):
    config = config
    UP_LIST = get_UP_list(config["sheetname"])
    if UP_LIST is None:
        logger.info(f'上传数据为空,请检查表格数据！')
        exit()
    else:
        logger.info(f'开始数据上传时间：{time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())}\n')
        # 遍历数据
        for k, v in UP_LIST.items():
            rawDataUrl, annotationsUrl, startTime, endTime = v
            # 打印开始时间
            logger.info(f'{time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())}：正在上传数据...{config["env"]},sheetname：{config["sheetname"]}  表格数据行号：{k} --> {v}\n')

            max_retries = 3
            for attempt in range(max_retries):
                try:
                    response = Sendto_webportal(rawDataUrl, annotationsUrl, startTime, endTime)
                    if response.status_code == 200:
                        logger.info(f'数据上传成功！{response.status_code} , {response.content}\n')
                        logger.info(f'{"*"*100}')
                        break
                    else:
                        raise Exception(f"数据上传出错！{response.status_code} , {response.content}\n")
                except Exception as e:
                    logger.info(f'数据上传出错！ 尝试 {attempt + 1}/{max_retries} {e}\n')
                    if attempt < max_retries - 1:
                        time.sleep(5)  # 等待5秒后重试
                    else:
                        logger.info(f'数据上传失败，已达到最大重试次数。\n')
                        logger.info(f'{"*"*100}')
        logger.info(f'完成数据上传时间：{time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())}\n')


# 读取配置信息
def get_config(Patient_json):
    # 读取配置信息 使用utf-8编码
    with open(Patient_json, "r", encoding="utf-8") as f:
        config = json.load(f)
    # print("配置信息：",config)
    return config

if __name__ == '__main__':

    Patient_json="patient_data_Dev.json"

    configs=get_config(Patient_json)
    for config in configs:
        main(config)
    # config = {
    #     "excle": "ecgDataChunks_66bda9b22acb763421454137_UAT.xlsx",
    # }
    #
    # lists=get_UP_list("ecgDataChunks")
    # for i in lists:
    #     print(i,lists[i])
