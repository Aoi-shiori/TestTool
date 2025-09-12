import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
# from .path_manager import DATA_PATH
from api_test_framework.utils.path_manager import DATA_PATH


class ExcelReader:
    def __init__(self, file_name="test_cases.xlsx"):
        # 首先尝试在子项目data目录中查找
        self.file_path = DATA_PATH / file_name
        print(self.file_path)

        # 如果不存在，尝试在项目根目录的data目录中查找
        if not self.file_path.exists():
            from .path_manager import PROJECT_ROOT
            root_data_path = PROJECT_ROOT / "data" / file_name
            print(root_data_path)
            if root_data_path.exists():
                self.file_path = root_data_path
            else:
                raise FileNotFoundError(f"Excel文件不存在: {self.file_path}")

    def read_test_cases(self, sheet_name=0):
        """读取Excel中的测试用例"""
        df = pd.read_excel(self.file_path, sheet_name=sheet_name)
        # 处理NaN值为空字符串
        df = df.fillna('')
        # 转换为字典列表
        test_cases = df.to_dict('records')
        return test_cases

    def write_test_results(self, results, sheet_name=0):
        """将测试结果写回Excel，并添加颜色标记"""
        # 使用openpyxl进行更精细的控制
        wb = load_workbook(self.file_path)

        # 获取工作表
        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]

        # 定义颜色填充
        pass_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 绿色
        fail_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 红色
        skip_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色

        # 找到测试结果列的索引
        header_row = 1
        result_col_idx = None
        for idx, cell in enumerate(ws[header_row], 1):
            if cell.value == "测试结果":
                result_col_idx = idx
                break

        if result_col_idx is None:
            # 如果没有测试结果列，添加一列
            max_col = ws.max_column
            ws.cell(row=header_row, column=max_col + 1, value="测试结果")
            result_col_idx = max_col + 1

        # 写入测试结果
        for i, result in enumerate(results, 2):  # 从第2行开始（第1行是标题）
            if i > ws.max_row:
                # 如果行数不够，添加新行
                ws.append([""] * ws.max_column)

            ws.cell(row=i, column=result_col_idx, value=result['result'])

            # 根据结果设置单元格颜色
            if result['result'] == "通过":
                ws.cell(row=i, column=result_col_idx).fill = pass_fill
            elif result['result'] == "失败":
                ws.cell(row=i, column=result_col_idx).fill = fail_fill
            elif result['result'] == "跳过":
                ws.cell(row=i, column=result_col_idx).fill = skip_fill

            # 如果有错误信息，写入备注列
            if 'error_msg' in result and result['error_msg']:
                # 查找备注列
                note_col_idx = None
                for idx, cell in enumerate(ws[header_row], 1):
                    if cell.value == "备注":
                        note_col_idx = idx
                        break

                if note_col_idx is None:
                    # 如果没有备注列，添加一列
                    max_col = ws.max_column
                    ws.cell(row=header_row, column=max_col + 1, value="备注")
                    note_col_idx = max_col + 1

                ws.cell(row=i, column=note_col_idx, value=result['error_msg'])

        # 保存文件
        wb.save(self.file_path)
if __name__ == '__main__':
    test = ExcelReader()
    data=test.read_test_cases()
    print(data)