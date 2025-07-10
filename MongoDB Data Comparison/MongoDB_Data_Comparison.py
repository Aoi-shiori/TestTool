# -*- coding: utf-8 -*-
"""
# @Creation time: 2025/7/3 14:12
# @Author       : 郭军
# @Email        : 391350540@qq.com
# @FileName     : MongoDB_Data_Comparison.py
# @Software     : PyCharm
# @Project      : MongoDB Data Comparison
# @PythonVersion: python 3.12
# @Version      : 
# @Description  : 
# @Update Time  : 
# @UpdateContent:  

"""
import math
from typing import Dict, Optional, Union, List, Any
# from bson import ObjectId
from pymongo import MongoClient, InsertOne, DeleteMany
import pandas as pd
from datetime import datetime
from urllib.parse import quote_plus
import time
from logger import logger
import gc
from openpyxl.styles import PatternFill
import openpyxl


def compare_collections(uri, db_name, collection1, collection2, output_file,query_date):
    """
    比较两个 MongoDB 集合并将结果写入 Excel 文件

    参数:
        uri: MongoDB 连接字符串
        db_name: 数据库名称
        collection1: 第一个集合名称
        collection2: 第二个集合名称
        output_file: 输出的 Excel 文件名
    """

    # 连接 MongoDB
    client = MongoClient(uri)
    db = client[db_name]

    # 获取两个集合
    coll1 = db[collection1]
    coll2 = db[collection2]

    # 获取集合中的所有文档
    docs1 = list(coll1.find(query_date))
    docs2 = list(coll2.find(query_date))

    # 将文档转换为 DataFrame
    df1 = pd.DataFrame(docs1)
    df2 = pd.DataFrame(docs2)

    # 删除 MongoDB 的 _id 字段，因为它每次都会不同
    # if '_id' in df1.columns:
    #     df1 = df1.drop('_id', axis=1)
    # if '_id' in df2.columns:
    #     df2 = df2.drop('_id', axis=1)

    # 比较两个 DataFrame
    # comparison = df1.compare(df2,keep_shape=True,keep_equal=True,align_axis=1)
    # comparison = df1.compare(df2,keep_shape=True,keep_equal=True,align_axis=1).reset_index()
    # comparison = df1.compare(df2,keep_equal=True,keep_shape=False,align_axis=0)
    # comparison = df1.compare(df2,keep_equal=True,align_axis=0)
    # comparison = df1.compare(df2,align_axis=1).reset_index()
    # comparison = df1.compare(df2,align_axis=0).reset_index()
    # comparison = df1.compare(df2).reset_index()
    # comparison = df1.compare(df2).fillna("PASS")
    comparison = df1.compare(df2).reset_index(drop=True)

    # 创建 Excel 写入对象
    writer = pd.ExcelWriter(output_file, engine='openpyxl')

    # 将原始数据和比较结果写入不同的工作表
    df1.to_excel(writer, sheet_name=collection1, index=False)
    df2.to_excel(writer, sheet_name=collection2, index=False)

    if not comparison.empty:
        comparison.to_excel(writer, sheet_name='Comparison', index=True)

        # 添加摘要信息
        summary = {
            'Comparison Date': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            'Collection 1': [collection1],
            'Collection 1 Count': [len(docs1)],
            'Collection 2': [collection2],
            'Collection 2 Count': [len(docs2)],
            'Differences Found': ['Yes'],
            'Difference Count': [len(comparison)]
        }
    else:
        summary = {
            'Comparison Date': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            'Collection 1': [collection1],
            'Collection 1 Count': [len(docs1)],
            'Collection 2': [collection2],
            'Collection 2 Count': [len(docs2)],
            'Differences Found': ['No']
        }

    pd.DataFrame(summary).to_excel(writer, sheet_name='Summary', index=False)

    # 保存 Excel 文件
    writer.close()
    logger.info(f"Comparison completed. Results saved to {output_file}")



# Excel行数限制 (Excel 2016及以后版本是1,048,576行)
EXCEL_ROW_LIMIT = 1000000  # 设置为略小于实际限制以留出空间

# 定义颜色填充
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')


def compare_large_collections(
        uri: str,
        db_name: str,
        collection1: str,
        collection2: str,
        output_prefix: str,
        query: Optional[Dict] = None,
        batch_size: Optional[int] = 1000,
        sample_size: int = 1000
) -> None:
    """
    比较大型MongoDB集合，自动处理Excel行数限制

    参数:
        uri: MongoDB连接字符串
        db_name: 数据库名称
        collection1: 第一个集合名称
        collection2: 第二个集合名称
        output_prefix: 输出Excel文件前缀(会自动添加序号)
        query: 可选查询条件(应用于两个集合)
        batch_size: 批量获取文档的大小
        sample_size: 每个集合保存的示例文档数
    """
    try:
        # 连接MongoDB
        client = MongoClient(uri)
        db = client[db_name]
        coll1 = db[collection1]
        coll2 = db[collection2]

        logger.info(f"开始比较大型集合: {collection1} 和 {collection2}")

        # 获取所有_id
        def get_ids(collection):
            return {doc['_id'] for doc in collection.find(query or {}, {'_id': 1}).batch_size(batch_size)}

        ids1 = get_ids(coll1)
        ids2 = get_ids(coll2)

        # 分类_id
        common_ids = list(ids1 & ids2)
        only_in_coll1 = list(ids1 - ids2)
        only_in_coll2 = list(ids2 - ids1)

        total_diffs = len(only_in_coll1) + len(only_in_coll2)
        logger.info(
            f"共有文档: {len(common_ids)}, 只在{collection1}中的文档: {len(only_in_coll1)}, 只在{collection2}中的文档: {len(only_in_coll2)}")

        # 计算需要的文件数量
        def calculate_file_count(diff_count):
            return max(1, math.ceil(diff_count / EXCEL_ROW_LIMIT))

        # 处理非共有文档的差异
        non_common_diffs = only_in_coll1 + only_in_coll2
        non_common_files = calculate_file_count(len(non_common_diffs))

        # 处理共有文档的字段差异
        field_diffs = []
        processed_ids = 0

        # 批量处理共有文档
        for i in range(0, len(common_ids), batch_size):
            batch_ids = common_ids[i:i + batch_size]

            # 批量获取文档
            # docs1 = {doc['_id']: doc for doc in coll1.find({'_id': {'$in': batch_ids}}, {'_id': 0})}
            # docs2 = {doc['_id']: doc for doc in coll2.find({'_id': {'$in': batch_ids}}, {'_id': 0})}
            docs1 = {doc['_id']: doc for doc in coll1.find({'_id': {'$in': batch_ids}})}
            docs2 = {doc['_id']: doc for doc in coll2.find({'_id': {'$in': batch_ids}})}

            for _id in batch_ids:
                doc1 = docs1.get(_id, {})
                doc2 = docs2.get(_id, {})

                # 获取所有字段时排除_id
                all_fields = set(doc1.keys()).union(set(doc2.keys())) - {'_id'}

                for field in all_fields:
                    val1 = doc1.get(field, '字段不存在')
                    val2 = doc2.get(field, '字段不存在')

                    if val1 != val2:
                        field_diffs.append({
                            '文档ID': _id,
                            '差异类型': '字段值不同',
                            '字段': field,
                            f'{collection1}值': str(val1),
                            f'{collection2}值': str(val2)
                        })

            processed_ids += len(batch_ids)
            if processed_ids % 10000 == 0:
                logger.info(f"已处理 {processed_ids}/{len(common_ids)} 个共有文档，发现 {len(field_diffs)} 处字段差异")

        total_diffs += len(field_diffs)
        field_diff_files = calculate_file_count(len(field_diffs))
        total_files = non_common_files + field_diff_files

        logger.info(f"总共发现 {total_diffs} 处差异，需要 {total_files} 个Excel文件")

        # 生成汇总报告文件
        generate_summary_file(
            output_prefix + "_summary.xlsx",
            collection1, collection2,
            len(ids1), len(ids2),
            len(common_ids), len(only_in_coll1), len(only_in_coll2),
            len(field_diffs), sample_size, coll1, coll2, query
        )

        # 生成差异文件 - 非共有文档
        generate_diff_files(
            output_prefix + "_non_common",
            non_common_files,
            only_in_coll1, only_in_coll2,
            collection1, collection2,
            coll1, coll2
        )

        # 生成差异文件 - 字段差异
        generate_diff_files(
            output_prefix + "_field_diffs",
            field_diff_files,
            diff_records=field_diffs,
            is_field_diff=True
        )

        logger.info("所有比较文件生成完成")

    except Exception as e:
        logger.error(f"比较过程中发生错误: {str(e)}", exc_info=True)
        raise
    finally:
        if 'client' in locals():
            client.close()


def generate_summary_file(
        file_path: str,
        coll1_name: str, coll2_name: str,
        count1: int, count2: int,
        common_count: int, only_in_coll1: int, only_in_coll2: int,
        field_diff_count: int,
        sample_size: int,
        coll1, coll2,
        query=None
) -> None:
    """生成汇总报告文件"""
    try:
        writer = pd.ExcelWriter(file_path, engine='openpyxl')

        # 1. 汇总工作表
        summary_data = {
            '比较时间': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            '集合1名称': [coll1_name],
            '集合1文档数': [count1],
            '集合2名称': [coll2_name],
            '集合2文档数': [count2],
            '共有文档数': [common_count],
            f'只在{coll1_name}中的文档数': [only_in_coll1],
            f'只在{coll2_name}中的文档数': [only_in_coll2],
            '字段差异数': [field_diff_count],
            '总差异数': [only_in_coll1 + only_in_coll2 + field_diff_count]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='汇总', index=False)

        # 2. 示例数据工作表
        def get_sample_docs(collection, size):
            cursor = collection.find(query or {}).limit(size)
            return list(cursor)

        sample_coll1 = get_sample_docs(coll1, sample_size)
        sample_coll2 = get_sample_docs(coll2, sample_size)

        pd.DataFrame(sample_coll1).to_excel(writer, sheet_name=f'{coll1_name}_集合', index=False)
        pd.DataFrame(sample_coll2).to_excel(writer, sheet_name=f'{coll2_name}_集合', index=False)

        writer.close()

        # 格式化Excel
        wb = openpyxl.load_workbook(file_path)
        ws = wb['汇总']

        # 设置列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter

            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        wb.save(file_path)

        logger.info(f"汇总报告已生成: {file_path}")

    except Exception as e:
        logger.error(f"生成汇总文件时出错: {str(e)}")


def generate_diff_files(
        file_prefix: str,
        file_count: int,
        only_in_coll1: Optional[List] = None,
        only_in_coll2: Optional[List] = None,
        coll1_name: Optional[str] = None,
        coll2_name: Optional[str] = None,
        coll1=None, coll2=None,
        diff_records: Optional[List] = None,
        is_field_diff: bool = False
) -> None:
    """生成差异文件"""
    if only_in_coll1 is not None and only_in_coll2 is not None:
        # 处理非共有文档
        all_non_common = []

        # 准备只在集合1中的文档记录
        for _id in only_in_coll1:
            doc = coll1.find_one({'_id': _id}, {'_id': 0})
            all_non_common.append({
                '文档ID': _id,
                '差异类型': f'只在{coll1_name}中存在',
                '字段': '全部',
                f'{coll1_name}值': str(doc),
                f'{coll2_name}值': '不存在'
            })

        # 准备只在集合2中的文档记录
        for _id in only_in_coll2:
            doc = coll2.find_one({'_id': _id}, {'_id': 0})
            all_non_common.append({
                '文档ID': _id,
                '差异类型': f'只在{coll2_name}中存在',
                '字段': '全部',
                f'{coll1_name}值': '不存在',
                f'{coll2_name}值': str(doc)
            })

        # 拆分到多个文件
        for i in range(file_count):
            start_idx = i * EXCEL_ROW_LIMIT
            end_idx = (i + 1) * EXCEL_ROW_LIMIT
            batch = all_non_common[start_idx:end_idx]

            if not batch:
                continue

            file_path = f"{file_prefix}_{i + 1}.xlsx"
            df = pd.DataFrame(batch)
            df.to_excel(file_path, index=False)
            apply_formatting(file_path, is_field_diff=False)

            logger.info(f"生成非共有文档差异文件: {file_path} (包含 {len(batch)} 条记录)")

    elif diff_records is not None:
        # 处理字段差异
        for i in range(file_count):
            start_idx = i * EXCEL_ROW_LIMIT
            end_idx = (i + 1) * EXCEL_ROW_LIMIT
            batch = diff_records[start_idx:end_idx]

            if not batch:
                continue

            file_path = f"{file_prefix}_{i + 1}.xlsx"
            df = pd.DataFrame(batch)
            df.to_excel(file_path, index=False)
            apply_formatting(file_path, is_field_diff=True)

            logger.info(f"生成字段差异文件: {file_path} (包含 {len(batch)} 条记录)")


def apply_formatting(file_path: str, is_field_diff: bool) -> None:
    """应用Excel格式化"""
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # 标题行样式
        for cell in ws[1]:
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        # 数据行样式
        for row in ws.iter_rows(min_row=2):
            if is_field_diff:
                # 字段差异使用黄色
                for cell in row:
                    cell.fill = YELLOW_FILL
            else:
                # 非共有文档根据类型着色
                diff_type = row[1].value  # 差异类型列
                if '存在' in diff_type:
                    fill_color = RED_FILL if '集合1' in diff_type else GREEN_FILL
                    for cell in row:
                        cell.fill = fill_color

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter

            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        wb.save(file_path)
    except Exception as e:
        logger.warning(f"无法应用Excel格式化: {str(e)}")


def collection_large_sharded_copy(mongodb_uri, source_db, source_col, target_db, target_col, batch_size, shard_key="_id"):
    """
    分片复制方案，避免内存溢出

    参数:
        shard_key: 用于分片的字段
        batch_size: 每批文档数
    """
    client = MongoClient(f'{mongodb_uri}',
                         connectTimeoutMS=3000000,
                         socketTimeoutMS=None,
                         maxPoolSize=15)

    source = client[source_db][source_col]
    target = client[target_db][target_col]
    logger.info("Mongodb链接创建完成！")

    # 清空目标集合
    try:
        logger.info(f"目标集合清理开始...")
        target.bulk_write([DeleteMany({})])
    except Exception as e:
        logger.error(f"清理目标集合出现异常：{e}")
    logger.info("目标集合数据清理完成！")
    # # 获取文档总数
    # total_docs = source.count_documents({})
    # logger.info(f"开始复制 {total_docs} 条文档...")

    # 1. 获取分片边界
    min_id = source.find_one(sort=[(shard_key, 1)])[shard_key]
    max_id = source.find_one(sort=[(shard_key, -1)])[shard_key]

    logger.info(f"{collection_a}分片复制范围: {min_id} 到 {max_id}")

    # 2. 定义分片大小(根据集合大小自动调整)
    total_docs = source.count_documents({})
    shard_count = max(10, total_docs // (batch_size * 100))  # 每10万文档一个分片
    logger.info(f"将分为 {shard_count} 个分片进行复制")

    # 3. 生成分片边界
    pipeline = [
        {"$bucketAuto": {
            "groupBy": f"${shard_key}",
            "buckets": shard_count,
            "output": {"min": {"$min": f"${shard_key}"}, "max": {"$max": f"${shard_key}"}}
        }}
    ]
    shards = list(source.aggregate(pipeline))

    # 4. 分片复制
    total_copied = 0
    for i, shard in enumerate(shards, 1):
        shard_min = shard["min"]
        shard_max = shard["max"]

        query = {shard_key: {"$gte": shard_min, "$lte": shard_max}}
        shard_docs = source.count_documents(query)

        logger.info(
            f"处理分片 {i}/{len(shards)}: {shard_key} 从 {shard_min} 到 {shard_max} "
            f"(约 {shard_docs} 文档)"
        )

        cursor = source.find(query).batch_size(batch_size)
        batch = []

        for doc in cursor:
            batch.append(InsertOne(doc))
            if len(batch) >= batch_size:
                target.bulk_write(batch, ordered=False)
                total_copied += len(batch)
                batch = []
                logger.debug(f"当前分片已复制: {total_copied}")

        if batch:
            target.bulk_write(batch, ordered=False)
            total_copied += len(batch)

        # 每个分片处理后释放内存
        del batch
        del cursor
        gc.collect()

    logger.info(f"{collection_a}复制完成! 共复制 {total_copied} 文档")



# 使用示例
if __name__ == "__main__":
    # MongoDB 连接配置
    name=quote_plus("jun")
    pwd=quote_plus("xsd@d234F66lk77@44fx")
    mongodb_uri = f"mongodb://{name}:{pwd}@localhost:2989/?connectTimeoutMS=19000000&authSource=webportal-dev&directConnection=true"
    # mongodb_uri = f"mongodb://{name}:{pwd}@webportal-k8s-dev-mongodb-0-f511ed4cc11a5904.elb.us-east-2.amazonaws.com:27017/?connectTimeoutMS=9000000&authSource=webportal-dev&directConnection=true"

    database_name = "webportal-dev2"  # 数据库名

    # 对比的集合名称
    # collection_a = "ecgEventChartData"  # 第一个集合名
    # collection_b = "ecgEventChartData_copy1"  # 第二个集合名

    collection_a = "ecgTraitData"  # 第一个集合名
    collection_b = "ecgTraitData_copy1"  # 第二个集合名

    # collection_a = "ecgBeatData"  # 第一个集合名
    # collection_b = "ecgBeatData_copy1"  # 第二个集合名

    # collection_a = "ecgEvents"  # 第一个集合名
    # collection_b = "ecgEvents_copy1"  # 第二个集合名

    # 查询参数
    clinic_ID="6848e8dd6b1fa7dec17a376e"

    output_excel = f"collection_comparison_{collection_a}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"  # 输出文件名


    #开始时间
    start=time.time()
    logger.info(f"开始时间：{datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S')}")

    # 数据集复制
    # collection_large_sharded_copy(mongodb_uri=mongodb_uri,source_db=database_name, source_col=collection_a, target_db=database_name, target_col=collection_b,batch_size=500)

    # # 数据集对比
    # compare_collections(
    #     uri=mongodb_uri,
    #     db_name=database_name,
    #     collection1=collection_a,
    #     collection2=collection_b,
    #     output_file=output_excel,
    #     # query_date = { "$and": [ { "createdAt": { "$gte": datetime(2021, 5, 29) } }, { "clinic": ObjectId(clinic_ID) } ] }
    #     # query_date = {"$and": [{"start": {"$lt": datetime(2025, 7, 7)}}]}
    #     query_date=None
    #
    #     # 时间参数
    #     # ecgBeatData: recordTime
    #     # ecgEvents: start
    #     # ecgTraitData: recordTime
    #     # ecgEventChartData: recordTime
    # )

    # 大型数据集对比
    compare_large_collections(
        uri=mongodb_uri,
        db_name=database_name,
        collection1=collection_a,
        collection2=collection_b,
        output_prefix=output_excel,
        # query={"status": "active"},  # 可选查询条件
        # query={"$and": [{"start": {"$lte": datetime(2025, 7, 7)}}]},  # 可选查询条件
        query={},  # 可选查询条件
        batch_size=5000,
        sample_size=1000
    )

    # 结束时间
    logger.info(f"结束时间：{datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"总计耗时：{time.time()-start:.2f}s")

