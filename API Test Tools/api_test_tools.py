#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
@Project: TestTool
@Author: guojun
@Email: 391350540@qq.com
@Date: 2026/4/7 13:33
@File: api_test_tools.py
@IDE: PyCharm
@Description: 
"""
import json
import re
import time
import pandas as pd
import requests
from openpyxl import load_workbook
from jsonschema import validate, ValidationError
from datetime import datetime
from urllib.parse import urlencode


class ExcelAPITester:
    def __init__(self, excel_path, sheet_name="API Test Result"):
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.test_results = []
        self.global_token = None
        self.base_url = "https://webportal-dev.vivalink.com/api/backend"  # 请根据实际环境修改

        # 默认必需的请求头
        self.default_headers = {
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
            'timezone': 'Asia/Shanghai',
            'dnt': '1',
            'priority': 'u=1, i',
            'sec-ch-ua': '"Chromium";v="146", "Not-A.Brand";v="24", "Microsoft Edge";v="146"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"macOS"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36 Edg/146.0.0.0'
        }

    def clean_json_string(self, json_str):
        """清理从Excel读取的JSON字符串"""
        if pd.isna(json_str) or not isinstance(json_str, str) or not json_str.strip():
            return None

        try:
            # 替换单引号为双引号
            json_str = json_str.replace("'", '"')
            # 移除JSON中的注释
            json_str = re.sub(r'//.*?\n|/\*.*?\*/', '', json_str, flags=re.DOTALL)
            # 去除非法控制字符
            json_str = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', json_str)
            return json.loads(json_str)
        except json.JSONDecodeError as e:
            print(f"JSON解析错误: {e}")
            return None

    def parse_query_params(self, params_str):
        """解析查询参数字符串为字典，支持JSON和键值对格式，并自动转换数字类型"""
        if pd.isna(params_str) or not isinstance(params_str, str) or not params_str.strip():
            return {}

        try:
            # 尝试作为JSON解析
            params_dict = self.clean_json_string(params_str)
            if isinstance(params_dict, dict):
                # 自动转换数字类型的字符串
                return self.convert_numeric_values(params_dict)
        except Exception as e:
            print(f"JSON解析查询参数失败，尝试键值对格式: {e}")

        # 如果不是JSON格式，尝试作为键值对解析
        params = {}
        pairs = params_str.strip().split('&')
        for pair in pairs:
            if '=' in pair:
                key, value = pair.split('=', 1)
                key = key.strip()
                value = value.strip()
                # 尝试转换数字值
                converted_value = self.convert_value_type(value)
                params[key] = converted_value

        print(f"解析后的查询参数: {params}")
        return params

    def convert_numeric_values(self, data):
        """递归转换字典中的数字字符串为数字类型"""
        if isinstance(data, dict):
            return {k: self.convert_numeric_values(v) for k, v in data.items()}
        elif isinstance(data, str):
            return self.convert_value_type(data)
        elif isinstance(data, list):
            return [self.convert_numeric_values(item) for item in data]
        return data

    def convert_value_type(self, value):
        """尝试将字符串转换为适当的类型（数字、布尔值等）"""
        if not isinstance(value, str):
            return value

        value = value.strip()

        # 尝试转换为整数
        if value.isdigit():
            return int(value)

        # 尝试转换为浮点数
        try:
            if '.' in value and value.replace('.', '', 1).replace('-', '', 1).isdigit():
                return float(value)
        except:
            pass

        # 尝试转换为布尔值
        if value.lower() == 'true':
            return True
        if value.lower() == 'false':
            return False

        # 尝试转换为null
        if value.lower() == 'null':
            return None

        return value

    def get_token(self, auth_row):
        """获取认证token"""
        print("正在获取认证token...")
        response = self.perform_request(auth_row, is_auth_request=True)

        if isinstance(response, str):
            print(f"获取token失败: {response}")
            return None

        try:
            response_json = response.json()
            # 从响应中提取token，根据实际API结构调整
            token = response_json.get('token') or response_json.get('accessToken') or response_json.get('data', {}).get(
                'token')
            if token:
                print(f"成功获取token: {token[:10]}...")
                return token
            else:
                print("未在响应中找到token字段")
                return None
        except Exception as e:
            print(f"解析token响应失败: {e}")
            return None

    def perform_request(self, row, is_auth_request=False):
        """执行HTTP请求，区分GET和POST参数"""
        path = row['接口路径']
        method = row['请求方法'].upper() if pd.notna(row['请求方法']) else 'GET'
        headers_str = row['请求头']
        body_str = row['请求体'] if method in ['POST', 'PUT', 'PATCH'] else None
        params_str = row['请求参数'] if method == 'GET' else None
        cookies_str = row.get('Cookies', '')  # 从Excel读取Cookies

        url = self.base_url + path

        # 合并默认头和自定义头
        headers = self.default_headers.copy()
        custom_headers = self.clean_json_string(headers_str) or {}
        headers.update(custom_headers)

        # 如果不是认证请求且有全局token，添加到headers
        if not is_auth_request and self.global_token:
            # 确保只设置一个认证头，避免冲突
            headers.pop('token', None)  # 移除可能存在的旧token字段
            headers['Authorization'] = f'Bearer {self.global_token}'

            # 设置referer头，根据接口路径动态生成
            if '/ecgReports' in path:
                headers['referer'] = f"{self.base_url.replace('/api/backend', '')}/report"
            elif '/patients' in path:
                headers['referer'] = f"{self.base_url.replace('/api/backend', '')}/patient"

        # 处理请求参数
        query_params = self.parse_query_params(params_str) if params_str else {}
        json_body = self.clean_json_string(body_str) if body_str else None

        # 处理cookies
        cookies = {}
        if cookies_str and isinstance(cookies_str, str):
            try:
                cookies = self.clean_json_string(cookies_str)
                if not isinstance(cookies, dict):
                    # 尝试解析cookie字符串
                    cookies = self.parse_cookie_string(cookies_str)
            except:
                cookies = self.parse_cookie_string(cookies_str)
        print(f"使用的Cookies: {cookies}")

        # 打印请求详情用于调试
        print(f"\n{'=' * 60}")
        print(f"📤 请求详情:")
        print(f"   URL: {url}")
        print(f"   Method: {method}")
        print(f"   Headers: {headers}")
        if method == 'GET' and query_params:
            print(f"   Query Params: {query_params}")
        elif json_body:
            print(f"   Request Body: {json_body}")
        print(f"   Cookies: {cookies}")
        print(f"{'=' * 60}")

        try:
            response = requests.request(
                method=method,
                url=url,
                headers=headers,
                params=query_params,  # GET请求参数
                json=json_body,  # POST/PUT请求体
                cookies=cookies,  # Cookies
                timeout=30
            )

            # 打印响应详情
            print(f"\n{'=' * 60}")
            print(f"📥 响应详情:")
            print(f"   Status Code: {response.status_code}")
            try:
                if response.status_code == 204:  # No Content
                    print(f"   Response Body: (无内容)")
                else:
                    response_json = response.json()
                    # 限制输出长度，避免太长
                    response_preview = json.dumps(response_json, indent=2, ensure_ascii=False)
                    if len(response_preview) > 500:
                        response_preview = response_preview[:500] + "... (内容过长)"
                    print(f"   Response Body: {response_preview}")
            except json.JSONDecodeError:
                response_preview = response.text[:500] + "..." if len(response.text) > 500 else response.text
                if response_preview.strip():  # 只有非空响应才显示
                    print(f"   Response Text: {response_preview}")
                else:
                    print(f"   Response: (空响应)")
            print(f"{'=' * 60}")

            return response
        except Exception as e:
            error_msg = f"请求异常: {str(e)}"
            print(f"❌ 请求异常: {error_msg}")
            return error_msg

    def parse_cookie_string(self, cookie_str):
        """解析cookie字符串为字典"""
        cookies = {}
        if not cookie_str or not isinstance(cookie_str, str):
            return cookies

        # 尝试不同的cookie格式
        if '; ' in cookie_str:
            # 格式: key1=value1; key2=value2
            pairs = cookie_str.split('; ')
            for pair in pairs:
                if '=' in pair:
                    key, value = pair.split('=', 1)
                    cookies[key.strip()] = value.strip()
        elif ',' in cookie_str:
            # 格式: key1=value1, key2=value2
            pairs = cookie_str.split(',')
            for pair in pairs:
                if '=' in pair:
                    key, value = pair.split('=', 1)
                    cookies[key.strip()] = value.strip()

        return cookies

    def validate_response(self, response, row):
        """验证响应结果"""
        case_id = row['用例ID']
        match_type = row['匹配方式'] if pd.notna(row['匹配方式']) else 'exact'
        expected_status = int(row['预期状态码']) if pd.notna(row['预期状态码']) else 200
        expected_response_str = row['预期响应']
        json_schema_str = row['JSON Schema']

        # 状态码验证
        if response.status_code != expected_status:
            return {
                "case_id": case_id,
                "result": "FAIL",
                "confirm_result": "FAIL",
                "summary": f"状态码不匹配: 预期{expected_status}, 实际{response.status_code}"
            }

        try:
            # 处理204 No Content
            if response.status_code == 204:
                actual_json = {}
            else:
                actual_json = response.json()
        except json.JSONDecodeError as e:
            actual_json = {}
            if match_type != "status_code":  # 如果不是只验证状态码，需要检查响应体
                return {
                    "case_id": case_id,
                    "result": "FAIL",
                    "confirm_result": "FAIL",
                    "summary": f"响应不是有效的JSON格式: {str(e)}"
                }
        except Exception as e:
            return {
                "case_id": case_id,
                "result": "FAIL",
                "confirm_result": "FAIL",
                "summary": f"解析响应失败: {str(e)}"
            }

        # 根据匹配类型进行验证
        if match_type == "status_code":
            return {
                "case_id": case_id,
                "result": "PASS",
                "confirm_result": "PASS",
                "summary": f"状态码匹配: {expected_status}"
            }

        elif match_type == "exact":
            expected_json = self.clean_json_string(expected_response_str)
            if expected_json is None:
                return {
                    "case_id": case_id,
                    "result": "FAIL",
                    "confirm_result": "FAIL",
                    "summary": "预期响应为空，无法进行完全匹配"
                }

            if expected_json == actual_json:
                return {
                    "case_id": case_id,
                    "result": "PASS",
                    "confirm_result": "PASS",
                    "summary": "完全匹配成功"
                }
            else:
                # 生成差异信息
                diff_info = self.get_diff_info(expected_json, actual_json)
                return {
                    "case_id": case_id,
                    "result": "FAIL",
                    "confirm_result": "FAIL",
                    "summary": f"完全匹配失败: {diff_info}"
                }

        elif match_type == "key_fields":
            expected_json = self.clean_json_string(expected_response_str)
            if not expected_json or not isinstance(expected_json, dict):
                return {
                    "case_id": case_id,
                    "result": "FAIL",
                    "confirm_result": "FAIL",
                    "summary": "预期响应格式错误，无法进行部分字段匹配"
                }

            match, missing_fields, mismatched_fields = self.partial_match(expected_json, actual_json)
            if match:
                return {
                    "case_id": case_id,
                    "result": "PASS",
                    "confirm_result": "PASS",
                    "summary": "部分字段匹配成功"
                }
            else:
                error_msg = []
                if missing_fields:
                    error_msg.append(f"缺失字段: {', '.join(missing_fields[:3])}")
                if mismatched_fields:
                    error_msg.append(f"值不匹配: {', '.join(mismatched_fields[:3])}")
                return {
                    "case_id": case_id,
                    "result": "FAIL",
                    "confirm_result": "FAIL",
                    "summary": "; ".join(error_msg)
                }

        elif match_type == "schema" and json_schema_str:
            schema = self.clean_json_string(json_schema_str)
            if not schema:
                return {
                    "case_id": case_id,
                    "result": "FAIL",
                    "confirm_result": "FAIL",
                    "summary": "JSON Schema为空，无法进行验证"
                }

            try:
                validate(instance=actual_json, schema=schema)
                return {
                    "case_id": case_id,
                    "result": "PASS",
                    "confirm_result": "PASS",
                    "summary": "Schema验证通过"
                }
            except ValidationError as e:
                return {
                    "case_id": case_id,
                    "result": "FAIL",
                    "confirm_result": "FAIL",
                    "summary": f"Schema验证失败: {e.message}"
                }

        # 默认情况 - 只要状态码正确就通过
        return {
            "case_id": case_id,
            "result": "PASS",
            "confirm_result": "PASS",
            "summary": "验证通过（默认）"
        }

    def partial_match(self, expected, actual, path=""):
        """部分字段匹配验证"""
        if not isinstance(expected, dict) or not isinstance(actual, dict):
            return expected == actual, [], []

        missing_fields = []
        mismatched_fields = []

        for key, exp_value in expected.items():
            current_path = f"{path}.{key}" if path else key

            if key not in actual:
                missing_fields.append(current_path)
                continue

            act_value = actual[key]

            if isinstance(exp_value, dict) and isinstance(act_value, dict):
                sub_match, sub_missing, sub_mismatched = self.partial_match(exp_value, act_value, current_path)
                if not sub_match:
                    missing_fields.extend(sub_missing)
                    mismatched_fields.extend(sub_mismatched)
            elif exp_value != act_value:
                mismatched_fields.append(f"{current_path}: 预期={exp_value}, 实际={act_value}")

        return len(missing_fields) == 0 and len(mismatched_fields) == 0, missing_fields, mismatched_fields

    def get_diff_info(self, expected, actual, max_items=3):
        """获取差异信息，限制显示数量"""
        if not isinstance(expected, dict) or not isinstance(actual, dict):
            return f"类型不匹配: 预期={type(expected)}, 实际={type(actual)}"

        diff_items = []

        # 检查缺失的字段
        for key in expected:
            if key not in actual:
                diff_items.append(f"缺失字段: {key}")

        # 检查多余的字段
        for key in actual:
            if key not in expected:
                diff_items.append(f"多余字段: {key}")

        # 检查值不匹配的字段
        for key in expected:
            if key in actual and expected[key] != actual[key]:
                # 限制值的显示长度
                exp_val = str(expected[key])[:50] + "..." if len(str(expected[key])) > 50 else str(expected[key])
                act_val = str(actual[key])[:50] + "..." if len(str(actual[key])) > 50 else str(actual[key])
                diff_items.append(f"值不匹配: {key} (预期={exp_val}, 实际={act_val})")

        # 限制显示数量
        if len(diff_items) > max_items:
            return "; ".join(diff_items[:max_items]) + f"... (共{len(diff_items)}处差异)"
        return "; ".join(diff_items[:max_items]) if diff_items else "无差异"

    def run_test_case(self, row, row_index):
        """运行单个测试用例"""
        case_id = row['用例ID']
        api_name = row['用例名称'] if pd.notna(row['用例名称']) else '未知接口'
        enabled = row['启用状态'] if pd.notna(row['启用状态']) else '是'

        if enabled != '是':
            print(f"\n⏭️ 跳过禁用的用例: {case_id} - {api_name}")
            return None

        print(f"\n{'=' * 60}")
        print(f"🧪 执行用例: {case_id} - {api_name}")
        print(f"🔧 接口路径: {row['接口路径']}")
        print(f"⚡ 请求方法: {row['请求方法'] if pd.notna(row['请求方法']) else 'GET'}")

        # 特殊处理认证接口
        interface_path = str(row['接口路径']).lower()
        if 'authentication' in interface_path or 'login' in interface_path or 'auth' in interface_path or 'signin' in interface_path:
            print("🔑 检测到认证接口，正在获取token...")
            token = self.get_token(row)
            if token:
                self.global_token = token
                result = {
                    "case_id": case_id,
                    "result": "PASS",
                    "confirm_result": "PASS",
                    "summary": "认证成功，token已获取并存储",
                    "row": row_index
                }
                print(f"✅ 认证成功，token已存储")
                return result
            else:
                result = {
                    "case_id": case_id,
                    "result": "FAIL",
                    "confirm_result": "FAIL",
                    "summary": "认证失败，无法获取token",
                    "row": row_index
                }
                print(f"❌ 认证失败，无法获取token")
                return result

        # 执行普通请求
        response = self.perform_request(row)

        if isinstance(response, str):  # 请求异常
            result = {
                "case_id": case_id,
                "result": "FAIL",
                "confirm_result": "FAIL",
                "summary": f"请求异常: {response}",
                "row": row_index
            }
            print(f"❌ 请求异常: {response}")
            return result

        # 验证响应
        validation_result = self.validate_response(response, row)
        validation_result["row"] = row_index

        status_icon = "✅" if validation_result['result'] == 'PASS' else "❌"
        print(f"\n{'=' * 40}")
        print(f"{status_icon} 验证结果: {validation_result['result']}")
        print(f"📝 详细信息: {validation_result['summary']}")
        print(f"{'=' * 40}")

        return validation_result

    def write_results_back(self):
        """将测试结果写回Excel文件"""
        print(f"\n{'=' * 60}")
        print("💾 正在将测试结果写入Excel文件...")

        try:
            # 加载工作簿
            wb = load_workbook(self.excel_path)
            if self.sheet_name not in wb.sheetnames:
                print(f"❌ 工作表 '{self.sheet_name}' 不存在")
                return False

            ws = wb[self.sheet_name]

            # 定义列索引（根据Excel列位置调整）
            # 假设：P列=确认结果(16), Q列=测试结果(17), S列=测试总结(19)
            CONFIRM_RESULT_COL = 16  # P列
            TEST_RESULT_COL = 17  # Q列
            SUMMARY_COL = 19  # S列
            EXECUTION_TIME_COL = 20  # T列，执行时间

            # 确保执行时间列有标题
            if ws.cell(row=1, column=EXECUTION_TIME_COL).value != "执行时间":
                ws.cell(row=1, column=EXECUTION_TIME_COL, value="执行时间")

            # 当前执行时间
            execution_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # 写入测试结果
            success_count = 0
            for result in self.test_results:
                if result is None:
                    continue

                row_idx = result['row'] + 2  # +2 因为Excel行号从1开始，且有标题行

                # 写入各列结果
                ws.cell(row=row_idx, column=TEST_RESULT_COL, value=result['result'])
                ws.cell(row=row_idx, column=CONFIRM_RESULT_COL, value=result['confirm_result'])
                ws.cell(row=row_idx, column=SUMMARY_COL, value=result['summary'])
                ws.cell(row=row_idx, column=EXECUTION_TIME_COL, value=execution_time)

                if result['result'] == 'PASS':
                    success_count += 1

            # 保存文件（覆盖原文件）
            wb.save(self.excel_path)
            print(f"✅ 测试结果已成功写入: {self.excel_path}")
            print(f"📊 写入结果: {success_count}/{len(self.test_results)} 个用例结果已更新")
            return True

        except Exception as e:
            print(f"❌ 写入Excel文件时出错: {e}")
            import traceback
            traceback.print_exc()
            return False

    def run(self):
        """主运行函数"""
        print(f"{'=' * 60}")
        print("🚀 开始执行API测试...")
        print(f"📁 Excel文件: {self.excel_path}")
        print(f"📋 工作表: {self.sheet_name}")
        print(f"🌐 基础URL: {self.base_url}")
        print(f"{'=' * 60}")

        try:
            # 读取Excel数据
            df = pd.read_excel(self.excel_path, sheet_name=self.sheet_name)
            print(f"📖 共读取 {len(df)} 条测试用例")

            # 检查Excel是否包含Cookies列，如果没有则添加
            if 'Cookies' not in df.columns:
                print("💡 Excel中缺少'Cookies'列，将使用默认cookie处理")
                df['Cookies'] = ''  # 添加空列

            # 按顺序执行测试用例
            for index, row in df.iterrows():
                result = self.run_test_case(row, index)
                if result is not None:
                    self.test_results.append(result)
                time.sleep(0.3)  # 避免请求过于频繁

            # 写回结果
            write_success = self.write_results_back()

            if not write_success:
                print("❌ 结果写入失败，但测试已执行完成")

            # 打印总结
            valid_results = [r for r in self.test_results if r is not None]
            total_cases = len(valid_results)
            passed_cases = len([r for r in valid_results if r['result'] == 'PASS'])
            failed_cases = total_cases - passed_cases

            print(f"\n{'=' * 60}")
            print("🎯 测试执行总结:")
            print(f"📝 总用例数: {total_cases}")
            print(f"✅ 通过用例: {passed_cases}")
            print(f"❌ 失败用例: {failed_cases}")
            if total_cases > 0:
                pass_rate = (passed_cases / total_cases) * 100
                print(f"📈 通过率: {pass_rate:.1f}%")
            print(f"{'=' * 60}")

            return {
                'total': total_cases,
                'passed': passed_cases,
                'failed': failed_cases,
                'success': write_success
            }

        except Exception as e:
            print(f"❌ 执行过程中发生错误: {e}")
            import traceback
            traceback.print_exc()
            return {
                'total': 0,
                'passed': 0,
                'failed': 0,
                'success': False,
                'error': str(e)
            }


if __name__ == "__main__":
    # 配置参数
    EXCEL_FILE = "./case/WebPortal API Test.xlsx"  # 请确保文件路径正确

    # 创建测试器并运行
    tester = ExcelAPITester(EXCEL_FILE)

    # 运行测试
    summary = tester.run()

    print(f"\n{'=' * 60}")
    print("🏁 测试执行完成！")
    if summary.get('success'):
        print("✅ 所有结果已成功写入原Excel文件")
    else:
        print("❌ 结果写入失败，请检查错误信息")
    print(f"{'=' * 60}")