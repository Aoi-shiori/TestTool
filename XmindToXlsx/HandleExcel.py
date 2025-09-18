#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
@Project: TestTool
@Author: guojun
@Email: 391350540@qq.com
@Date: 2025/9/17 11:58
@File: HandleExcel.py
@IDE: PyCharm
@Description: 
"""

from logger import logger as logger
import re
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import io
import math


class HandleExcel():
    def __init__(self, fileName, filePath):
        self.fileName = fileName
        self.workbook = Workbook()
        self.worksheet = None
        self.case_demo = {}
        self.filePath = filePath

        # 图片单元格的尺寸（单位：像素） - 增加尺寸以提高清晰度
        self.image_cell_width = 400  # 增加宽度以提高清晰度
        self.image_cell_height = 300  # 增加高度以提高清晰度

        # 设置图片质量参数
        self.image_quality = 100  # 图片质量（0-100）
        self.max_image_size = 1920  # 最大图片尺寸（长边）

    def _resize_image(self, unique_id, image_data):
        """高质量调整图片大小，保持清晰度"""
        try:
            # 将字节数据转换为PIL图像
            img = PILImage.open(io.BytesIO(image_data))

            # 获取原始尺寸和格式
            width, height = img.size
            original_format = img.format or 'PNG'

            # 计算缩放比例，保持宽高比
            # 只有当图片尺寸超过最大限制时才进行缩放
            max_dimension = max(width, height)
            if max_dimension > self.max_image_size:
                scale_ratio = self.max_image_size / max_dimension
                new_width = int(width * scale_ratio)
                new_height = int(height * scale_ratio)

                # 使用高质量的重采样算法
                img = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
            else:
                # 如果图片小于最大限制，保持原始尺寸
                new_width = width
                new_height = height

            # 保存到内存中，保持高质量
            img_byte_arr = io.BytesIO()

            # 根据格式选择保存方法
            if original_format.upper() in ['JPEG', 'JPG']:
                # 对于JPEG，使用高质量设置
                img.save(img_byte_arr, format='JPEG', quality=self.image_quality, optimize=True)
            else:
                # 对于其他格式，使用PNG以保持质量
                img.save(img_byte_arr, format='PNG', optimize=True)

            img_byte_arr.seek(0)

            # 创建Excel图片对象
            excel_img = XLImage(img_byte_arr)

            # 确保图片不会超过单元格大小
            if excel_img.width > self.image_cell_width or excel_img.height > self.image_cell_height:
                # 计算缩放比例以适应单元格
                width_ratio = self.image_cell_width / excel_img.width
                height_ratio = self.image_cell_height / excel_img.height
                ratio = min(width_ratio, height_ratio)

                # 调整Excel图片对象的尺寸
                excel_img.width = int(excel_img.width * ratio)
                excel_img.height = int(excel_img.height * ratio)

            return excel_img, excel_img.width, excel_img.height

        except Exception as e:
            logger.error(f"调整用例 ID：{unique_id}图片大小失败: {str(e)}，请检查图片是否正确！！")
            return None, 0, 0

    def _set_cell_size(self, worksheet, row, col, width, height):
        """设置单元格大小以适应图片"""
        # 设置列宽 (1个字符宽度 ≈ 7像素)
        col_letter = get_column_letter(col)
        char_width = max(10, math.ceil(width / 7))
        worksheet.column_dimensions[col_letter].width = min(char_width, 50)  # 限制最大列宽

        # 设置行高 (1磅 ≈ 1.33像素)
        point_height = max(15, math.ceil(height / 1.33))
        worksheet.row_dimensions[row].height = min(point_height, 400)  # 限制最大行高

    def generate_title(self, maxModule, data_list):
        status = ""
        data_list = data_list.copy()

        title_list = ["Number"]
        # 根据模块的级数生成对应的标题
        for item in range(maxModule):
            module = "Module-" + str(item + 1)
            title_list.append(module)

        # 适配Webportal用例逻辑
        datalist_lenth = len(data_list)
        if datalist_lenth > 1:
            lenth = 1
        else:
            lenth = 0

        if maxModule == 5 and "PRD" in data_list[lenth]["Case"][0].get("module-5", ""):
            title_list = ["Number", "Version", "PRD", "Module-1", "Module-2", "Module-3"]
        elif maxModule == 6 and "PRD" in data_list[lenth]["Case"][0].get("module-5", ""):
            title_list = ["Number", "Version", "PRD", "Module-1", "Module-2", "Module-3", "Module-6"]
        else:
            pass

        if "Version" in title_list and "PRD" in title_list:
            # 重组数据列名
            key_mapping = {"module-4": "version", "module-5": "prd"}  # 定义键名映射
            for data in data_list:
                case_list = data["Case"]
                new_case_list = []
                for case in case_list:
                    case = {key_mapping.get(k, k): v for k, v in case.items()}
                    new_case_list.append(case)
                data["Case"] = new_case_list
        else:
            pass

        title_list += ["Test Item", "Preconditions", "Test Step", "Expected Result", "Result", "Regression", "Note",
                       "Status", "验证图片1", "验证图片2", "验证图片3", "验证图片4", "验证图片5", "验证图片6",
                       "验证图片7", "验证图片8", "验证图片9", "验证图片10"]

        # 获取title名称并创建sheet
        for data in data_list:
            sheetname = data['title']
            logger.info(f'sheetname:,{sheetname}')

            # 删除默认的Sheet（如果存在）
            if "Sheet" in self.workbook.sheetnames:
                std_sheet = self.workbook["Sheet"]
                self.workbook.remove(std_sheet)

            self.worksheet = self.workbook.create_sheet(sheetname)

            # 定义样式
            header_font = Font(bold=True, size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            alignment = Alignment(vertical='top', wrap_text=True)

            # 将生成的标题写入到excel中
            for i in range(len(title_list)):
                cell = self.worksheet.cell(row=1, column=i + 1, value=title_list[i])
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # 存储列号映射
                if title_list[i] == "Test Item":
                    self.case_demo['title'] = i + 1
                elif title_list[i] == "Test Step":
                    self.case_demo['TestStep'] = i + 1
                elif title_list[i] == "Expected Result":
                    self.case_demo['ExpectedResult'] = i + 1
                elif title_list[i] == "Result":
                    self.case_demo['case_status'] = i + 1
                elif title_list[i] == "Regression":
                    self.case_demo['regression'] = i + 1
                elif title_list[i] == "Note":
                    self.case_demo['note'] = i + 1
                elif title_list[i] == "Status":
                    self.case_demo['status'] = i + 1
                elif title_list[i] == "验证图片1":
                    self.case_demo['image1'] = i + 1
                elif title_list[i] == "验证图片2":
                    self.case_demo['image2'] = i + 1
                elif title_list[i] == "验证图片3":
                    self.case_demo['image3'] = i + 1
                elif title_list[i] == "验证图片4":
                    self.case_demo['image4'] = i + 1
                elif title_list[i] == "验证图片5":
                    self.case_demo['image5'] = i + 1
                elif title_list[i] == "验证图片6":
                    self.case_demo['image6'] = i + 1
                elif title_list[i] == "验证图片7":
                    self.case_demo['image7'] = i + 1
                elif title_list[i] == "验证图片8":
                    self.case_demo['image8'] = i + 1
                elif title_list[i] == "验证图片9":
                    self.case_demo['image9'] = i + 1
                elif title_list[i] == "验证图片10":
                    self.case_demo['image10'] = i + 1
                else:
                    self.case_demo[title_list[i].lower()] = i + 1

            # 设置列宽，特别是图片列的宽度
            column_widths = [8, 15, 15, 15, 15, 30, 20, 30, 30, 10, 10, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20]
            for i, width in enumerate(column_widths[:len(title_list)]):
                self.worksheet.column_dimensions[get_column_letter(i + 1)].width = width

            case_list = data['Case']
            status = self.write_data(case_list)

        # 保存文件
        try:
            self.workbook.save(f"{self.filePath}.xlsx")
            return True
        except Exception as e:
            logger.error(f"保存Excel文件失败: {str(e)}")
            return False

    def write_data(self, data_list):
        logger.info(f'用例数量：{len(data_list)}')

        try:
            for item in range(len(data_list)):
                row = item + 2  # 从第二行开始（第一行是标题）
                case = data_list[item]
                unique_id = item + 1  # 默认用例ID

                # 写入用例ID
                try:
                    Note = case.get('note', '')
                    if Note and Note != '':
                        match = re.search(r'(\d{4})(\d{2})(\d{2})(\d{2})(\d{2})(\d{2})', Note)
                        if match:
                            year, month, day, hour, minute, second = match.groups()
                            date_id = f'{year}{month}{day}{hour}{minute}{second}'
                            unique_id = int(date_id)
                            self.worksheet.cell(row=row, column=1, value=unique_id)
                        else:
                            self.worksheet.cell(row=row, column=1, value=unique_id)
                    else:
                        self.worksheet.cell(row=row, column=1, value=unique_id)
                except:
                    self.worksheet.cell(row=row, column=1, value=unique_id)

                # 写入其他数据
                for key, value in case.items():
                    if key in self.case_demo and 'image_data' not in key:
                        col = self.case_demo[key]
                        cell = self.worksheet.cell(row=row, column=col, value=str(value))

                        # 设置样式
                        cell.alignment = Alignment(vertical='top', wrap_text=True)

                        # 根据执行结果设置背景色
                        if key == 'case_status':
                            if value == 'PASS':
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif value == 'FAIL':
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                # 插入图片
                for i in range(1, 11):
                    image_key = f"image{i}"
                    image_data_key = f"image_data{i}"

                    if image_key in self.case_demo and image_data_key in case:
                        image_col = self.case_demo[image_key]
                        try:
                            img, img_width, img_height = self._resize_image(unique_id, case[image_data_key])
                            if img:
                                # 设置单元格大小以适应图片
                                self._set_cell_size(self.worksheet, row, image_col, img_width, img_height)

                                # 将图片锚定到单元格
                                cell_addr = f'{get_column_letter(image_col)}{row}'
                                img.anchor = cell_addr
                                self.worksheet.add_image(img)
                        except Exception as e:
                            logger.error(f"插入图片失败: {str(e)}")
                            self.worksheet.cell(row=row, column=image_col, value="图片加载失败")
                    elif image_key in self.case_demo:
                        image_col = self.case_demo[image_key]
                        self.worksheet.cell(row=row, column=image_col, value="无图片")

            return True
        except Exception as e:
            logger.error(f"写入数据失败: {str(e)}")
            return False